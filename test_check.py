import openpyxl
import logging
from phone import Phone
from base.base_path import ReadPath


class CheckData:

    # 打开当前已有数据excel表
    excel = openpyxl.load_workbook(ReadPath().excel_path())

    # 实例化才能对这个表进行操作
    table = excel.active

    # 获取行和列
    rows = table.max_row
    cols = table.max_column

    print(rows, cols)

    def get_phone(self):

        for i in range(1, self.rows):
            phone_num = self.table.cell(i + 1, self.cols).value
            #print(phone_num)
            info = Phone().find(phone_num)
            # print(info)
            # print(info['phone'])
            # print(info['phone_type'])
            logging.warning('手机号:{x}, 归属地查询:{y}'.format(x=info['phone'], y=info['phone_type']))


check = CheckData()
check.get_phone()
